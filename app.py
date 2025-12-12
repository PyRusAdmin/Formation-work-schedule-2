# -*- coding: utf-8 -*-
import calendar
import json
import secrets
from datetime import date, datetime
from typing import Annotated
from typing import List

from fastapi import FastAPI, Request, Depends, HTTPException, Form, Path
from fastapi.responses import HTMLResponse
from fastapi.responses import Response
from fastapi.security import HTTPBasic, HTTPBasicCredentials
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from loguru import logger  # https://github.com/Delgan/loguru
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from peewee import IntegrityError
from pydantic import BaseModel

from config import AUTHORIZED_USERNAME, AUTHORIZED_PASSWORD
from database import initialize_db, ReportCard10, ReportCard11, ReportCard12, ReportCard01, DataStaff, db

logger.add("log/log.log")  # 📝 логирование

security = HTTPBasic()  # ← это обязательный объект

app = FastAPI()  # Создаем экземпляр FastAPI
# Монтируем статические файлы
app.mount('/static', StaticFiles(directory='static'), name='static')
# Создаем экземпляр Jinja2Templates
templates = Jinja2Templates(directory="templates")

# Инициализация БД при запуске приложения
initialize_db()


# Модели Pydantic
class EmployeeCreate(BaseModel):
    service_number: str
    vacation_start: date
    vacation_end: date


class EmployeeResponse(BaseModel):
    id: int
    service_number: str
    vacation_start: date
    vacation_end: date


@app.get("/data_10")
async def get_data():
    """Получение данных из БД октябрь 2025 года"""
    employees = []
    for emp in ReportCard10.select():
        employees.append({
            "КСП": emp.ksp,
            "Наименование": emp.name,
            "Категория": emp.category,
            "Профессия": emp.profession,
            "Статус": emp.status,
            "Сокращение": emp.abbreviation,
            "Разряд": emp.grade,
            "Таб": emp.tab,
            "ФИО": emp.fio,
            "Тариф": emp.salary,
            "days": json.loads(emp.days)
        })
    return employees


@app.post("/data_10")
async def save_data(request: Request):
    """Сохранение данных в БД и запись даты изменения октябрь 2025 года"""
    new_data = await request.json()
    now = datetime.now()  # текущее время

    for row in new_data:
        emp, created = ReportCard10.get_or_create(tab=row["Таб"])
        emp.ksp = row["КСП"]
        emp.name = row["Наименование"]
        emp.category = row["Категория"]
        emp.profession = row["Профессия"]
        emp.status = row["Статус"]
        emp.abbreviation = row.get("Сокращение", "")
        emp.grade = row.get("Разряд", "")
        emp.fio = row["ФИО"]
        emp.salary = row["Тариф"]
        emp.days = json.dumps(row["days"], ensure_ascii=False)
        emp.date_change = now  # 🕒 записываем текущие дату и время
        emp.save()
    return {"status": "ok"}


@app.get("/data_11")
async def get_data():
    """Получение данных из БД ноябрь 2025 года"""
    employees = []
    for emp in ReportCard11.select():
        employees.append({
            "КСП": emp.ksp,
            "Наименование": emp.name,
            "Категория": emp.category,
            "Профессия": emp.profession,
            "Статус": emp.status,
            "Сокращение": emp.abbreviation,
            "Разряд": emp.grade,
            "Таб": emp.tab,
            "ФИО": emp.fio,
            "Тариф": emp.salary,
            "days": json.loads(emp.days)
        })
    return employees


@app.post("/data_11")
async def save_data(request: Request):
    """Сохранение данных в БД и запись даты изменения ноябрь 2025 года"""
    new_data = await request.json()
    now = datetime.now()  # текущее время

    for row in new_data:
        emp, created = ReportCard11.get_or_create(tab=row["Таб"])
        emp.ksp = row["КСП"]
        emp.name = row["Наименование"]
        emp.category = row["Категория"]
        emp.profession = row["Профессия"]
        emp.status = row["Статус"]
        emp.abbreviation = row.get("Сокращение", "")
        emp.grade = row.get("Разряд", "")
        emp.fio = row["ФИО"]
        emp.salary = row["Тариф"]
        emp.days = json.dumps(row["days"], ensure_ascii=False)
        emp.date_change = now  # 🕒 записываем текущие дату и время
        emp.save()
    return {"status": "ok"}


@app.get("/data_12")
async def get_data():
    """Получение данных из БД декабрь 2025 года"""
    employees = []
    for emp in ReportCard12.select():
        employees.append({
            "КСП": emp.ksp,
            "Наименование": emp.name,
            "Категория": emp.category,
            "Профессия": emp.profession,
            "Статус": emp.status,
            "Сокращение": emp.abbreviation,
            "Разряд": emp.grade,
            "Таб": emp.tab,
            "ФИО": emp.fio,
            "Тариф": emp.salary,
            "days": json.loads(emp.days)
        })
    return employees


@app.post("/data_12")
async def save_data(request: Request):
    """Сохранение данных в БД и запись даты изменения декабрь 2025 года"""
    new_data = await request.json()
    now = datetime.now()  # текущее время

    for row in new_data:
        emp, created = ReportCard12.get_or_create(tab=row["Таб"])
        emp.ksp = row["КСП"]
        emp.name = row["Наименование"]
        emp.category = row["Категория"]
        emp.profession = row["Профессия"]
        emp.status = row["Статус"]
        emp.abbreviation = row.get("Сокращение", "")
        emp.grade = row.get("Разряд", "")
        emp.fio = row["ФИО"]
        emp.salary = row["Тариф"]
        emp.days = json.dumps(row["days"], ensure_ascii=False)
        emp.date_change = now  # 🕒 записываем текущие дату и время
        emp.save()
    return {"status": "ok"}


@app.get("/report_card_10", response_model=None)
async def report_card_10(request: Request):
    """
    Страница формирования табеля сотрудников октябрь 2025 года
    """
    return templates.TemplateResponse("work_schedule/2025/10/report_card_10.html", {"request": request})


@app.get("/report_card_11", response_model=None)
async def report_card_11(request: Request):
    """
    Страница формирования табеля сотрудников ноябрь 2025 года
    """
    return templates.TemplateResponse("work_schedule/2025/11/report_card_11.html", {"request": request})


@app.get("/report_card_12", response_model=None)
async def report_card_12(request: Request):
    """
    Страница формирования табеля сотрудников декабрь 2025 года
    :param request: Request - запрос
    :return: HTMLResponse - ответ
    """
    try:
        return templates.TemplateResponse("work_schedule/2025/12/report_card_12.html", {"request": request})
    except Exception as e:
        logger.exception(e)


@app.get("/download_excel_12")
async def download_excel_12():
    """
    Эндпоинт для скачивания Excel-файла с данными за декабрь 2025
    :return: Response - файл Excel
    """
    # Получаем данные из базы
    employees = []
    for emp in ReportCard12.select():
        employees.append({
            "КСП": emp.ksp,
            "Наименование": emp.name,
            "Категория": emp.category,
            "Профессия": emp.profession,
            "Статус": emp.status,
            "Таб": emp.tab,
            "ФИО": emp.fio,
            "Тариф": emp.salary,
            "days": json.loads(emp.days)
        })

    # Создаем Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "График декабрь 2025"

    # Заголовки
    headers = ["КСП", "Наименование", "Категория", "Профессия", "Статус", "Таб", "ФИО", "Тариф"] + [str(i) for i in range(1, 32)]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Данные
    for row_idx, emp in enumerate(employees, 2):
        # Основные данные
        ws.cell(row=row_idx, column=1).value = emp["КСП"]
        ws.cell(row=row_idx, column=2).value = emp["Наименование"]
        ws.cell(row=row_idx, column=3).value = emp["Категория"]
        ws.cell(row=row_idx, column=4).value = emp["Профессия"]
        ws.cell(row=row_idx, column=5).value = emp["Статус"]
        ws.cell(row=row_idx, column=6).value = emp["Таб"]
        ws.cell(row=row_idx, column=7).value = emp["ФИО"]
        ws.cell(row=row_idx, column=8).value = emp["Тариф"]

        # Дни месяца
        for day_idx, day_value in enumerate(emp["days"]):
            cell = ws.cell(row=row_idx, column=9 + day_idx)
            cell.value = day_value
            cell.alignment = Alignment(horizontal="center", vertical="center")
            # Цвета
            if day_value == "Б":
                cell.fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
            elif day_value == "О":
                cell.fill = PatternFill(start_color="C5CAE9", end_color="C5CAE9", fill_type="solid")
            elif day_value == "1":
                cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
            elif day_value == "ПС":
                cell.fill = PatternFill(start_color="FFECB3", end_color="FFECB3", fill_type="solid")
            elif day_value == "ДО":
                cell.fill = PatternFill(start_color="D1C4E9", end_color="D1C4E9", fill_type="solid")
            elif day_value == "БД":
                cell.fill = PatternFill(start_color="B3E5FC", end_color="B3E5FC", fill_type="solid")
            elif day_value == "Г":
                cell.fill = PatternFill(start_color="B2DFDB", end_color="B2DFDB", fill_type="solid")
            elif day_value in ["В", "в"]:
                cell.fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
            elif day_value == "-":
                cell.fill = PatternFill(start_color="CFD8DC", end_color="CFD8DC", fill_type="solid")

    # Автоширина
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 20)
        ws.column_dimensions[column].width = adjusted_width

    # Высота строк
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        ws.row_dimensions[row[0].row].height = 20

    # Ответ
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    headers = {
        "Content-Disposition": "attachment; filename=report_card_12.xlsx"
    }
    return Response(content=output.read(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers=headers)


@app.get("/report_card_01", response_model=None)
async def report_card_12(request: Request):
    """
    Страница формирования табеля сотрудников январь 2026 года
    :param request: Request - запрос
    :return: HTMLResponse - ответ
    """
    try:
        return templates.TemplateResponse("work_schedule/2026/01/report_card_01.html", {"request": request})
    except Exception as e:
        logger.exception(e)


def authenticate_user(credentials: HTTPBasicCredentials = Depends(security)):
    # Защита от None (если переменные не заданы в .env)
    expected_username = AUTHORIZED_USERNAME or ""
    expected_password = AUTHORIZED_PASSWORD or ""

    correct_username = secrets.compare_digest(credentials.username, expected_username)
    correct_password = secrets.compare_digest(credentials.password, expected_password)

    if not (correct_username and correct_password):
        raise HTTPException(
            status_code=401,
            detail="Неверный логин или пароль",
            headers={"WWW-Authenticate": "Basic"},
        )
    return credentials.username


@app.get("/list_employees", response_class=HTMLResponse)
async def list_employees(
        request: Request,
        _username: str = Depends(authenticate_user)  # ← добавили зависимость
):
    """
    Страница списка сотрудников (требует авторизации)
    """
    try:
        employees = []
        for emp in ReportCard10.select():
            date_change = emp.date_change.strftime("%d.%m.%Y %H:%M") if emp.date_change else "—"
            employees.append({
                "ksp": emp.ksp,
                "name": emp.name,
                "category": emp.category,
                "profession": emp.profession,
                "status": emp.status,
                "abbreviation": emp.abbreviation,
                "grade": emp.grade,
                "tab": emp.tab,
                "fio": emp.fio,
                "salary": emp.salary,
                "date_change": date_change,
            })

        return templates.TemplateResponse(
            "list_employees.html",
            {"request": request, "employees": employees}
        )
    except Exception as e:
        logger.exception(e)
        return {"error": str(e)}


# CRUD операции
@app.post("/employees/", response_model=EmployeeResponse)
async def create_employee(employee: EmployeeCreate):
    new_employee = ReportCard10.create(
        name=employee.service_number,
        vacation_start=employee.vacation_start,
        vacation_end=employee.vacation_end,
    )
    return EmployeeResponse(
        id=new_employee.id,
        service_number=new_employee.service_number,
        vacation_start=new_employee.vacation_start,
        vacation_end=new_employee.vacation_end,
    )


@app.get("/employees/", response_model=List[EmployeeResponse])
async def get_employees():
    employees = ReportCard10.select()
    return [
        EmployeeResponse(
            id=emp.id,
            service_number=emp.service_number,
            vacation_start=emp.vacation_start,
            vacation_end=emp.vacation_end,
        )
        for emp in employees
    ]


@app.get("/entering_vacations", response_model=None)
async def entering_vacations(request: Request):
    """
    Страница ввода отпусков
    """
    return templates.TemplateResponse("entering_vacations.html", {"request": request})


@app.get("/calendar_2025", response_model=None)
async def calendar_2025(request: Request):
    """
    Страница календаря 2025 года
    :param request: FastAPI request
    :return: templates.TemplateResponse
    """
    return templates.TemplateResponse("choosing_month.html", {"request": request})


"""Формирование графика по сотруднику"""


@app.get("/forming_employee_report_card_12", response_model=None)
async def forming_employee_report_card_12(request: Request):
    """
    Формирование графика сотрудника (декабрь 2025)
    :param request: FastAPI request
    :return: templates.TemplateResponse
    """
    return templates.TemplateResponse("work_schedule/forming_employee_report_card_12.html", {"request": request})


@app.get("/forming_employee_report_card", response_model=None)
async def forming_employee_report_card(request: Request):
    """
    Формирование графика сотрудника (ноябрь 2025)
    :param request: FastAPI request
    :return: templates.TemplateResponse
    """
    return templates.TemplateResponse("work_schedule/forming_employee_report_card.html", {"request": request})


"""Формирование личного дела по сотруднику"""


@app.get(path="/personal_business", response_model=None)
async def personal_business(request: Request, message: str = None):
    """Формирование личного дела по сотруднику"""
    return templates.TemplateResponse("personal_business.html", {"request": request, "message": message})


@app.post("/delete/")
async def delete_employee(
        request: Request,
        service_number: Annotated[str, Form()],
        dismissal_date: Annotated[str, Form()],  # ← принимаем как строку
        month: Annotated[str, Form()]
):
    logger.info(f"Табельный номер {service_number}, дата сокращения {dismissal_date}, месяц {month}")
    try:
        # Преобразуем строку в date
        dismissal_date_obj = datetime.strptime(dismissal_date, "%Y-%m-%d").date()
    except ValueError:
        message = "⚠️ Неверный формат даты. Ожидается ГГГГ-ММ-ДД."
        return templates.TemplateResponse("personal_business.html", {"request": request, "message": message})

    message = ""

    with db.atomic():
        emp = DataStaff.get_or_none(DataStaff.service_number == service_number)
        if emp:
            emp.dismissal_date = dismissal_date_obj  # ← используем объект date
            emp.save()
            message += f"✅ Сотрудник {emp.person} ({service_number}) уволен {dismissal_date_obj}.<br>"
        else:
            message = f"⚠️ Сотрудник с табельным номером {service_number} не найден."
            return templates.TemplateResponse("personal_business.html", {"request": request, "message": message})

        if int(month) <= 10:
            ReportCard10.delete().where(ReportCard10.tab == service_number).execute()
        if int(month) <= 11:
            ReportCard11.delete().where(ReportCard11.tab == service_number).execute()
        if int(month) <= 12:
            ReportCard12.delete().where(ReportCard12.tab == service_number).execute()

        message += f"🧹 Удалён из графиков, начиная с месяца №{month}."

    return templates.TemplateResponse("personal_business.html", {"request": request, "message": message})


# === Эндпоинты для формирования графика по табельному номеру (ноябрь 2025 → ReportCard11) ===


@app.get("/api/employee/{tab}")
async def get_employee_by_tab(tab: str):
    """Получить сотрудника по табельному номеру из ноября 2025 (ReportCard11)"""
    try:
        emp = ReportCard11.get(ReportCard11.tab == tab)
        return {
            "id": emp.id,
            "tab": emp.tab,
            "fio": emp.fio,
            "ksp": emp.ksp,
            "name": emp.name,
            "category": emp.category,
            "profession": emp.profession,
            "status": emp.status,
            "abbreviation": emp.abbreviation,
            "grade": emp.grade,
            "salary": emp.salary,
            "days": json.loads(emp.days)
        }
    except ReportCard11.DoesNotExist:
        raise HTTPException(status_code=404, detail="Сотрудник с таким табельным номером не найден в ноябре 2025")


@app.put("/api/employee/{tab}")
async def update_employee_days(tab: str, request: Request):
    """Обновить график сотрудника (только days и date_change)"""
    try:
        emp = ReportCard11.get(ReportCard11.tab == tab)
    except ReportCard11.DoesNotExist:
        raise HTTPException(status_code=404, detail="Сотрудник не найден")

    try:
        data = await request.json()
        new_days = data.get("days")

        if not isinstance(new_days, list):
            raise HTTPException(status_code=400, detail="Поле 'days' должно быть списком")

        if len(new_days) != 30:
            raise HTTPException(status_code=400,
                                detail="Ноябрь 2025 имеет 30 дней. Передано: {} дней".format(len(new_days)))

        # Обновляем только days и date_change
        emp.days = json.dumps(new_days, ensure_ascii=False)
        emp.date_change = datetime.now()
        emp.save()

        return {"status": "ok", "message": "График успешно обновлён"}
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Некорректный JSON")
    except Exception as e:
        logger.exception(e)
        raise HTTPException(status_code=500, detail="Ошибка при сохранении")


# === Эндпоинты для формирования графика по табельному номеру (декабрь 2025 → ReportCard12) ===

# Словарь для сопоставления месяца и модели
REPORT_CARD_MODELS = {
    10: ReportCard10,
    11: ReportCard11,
    12: ReportCard12,
    1: ReportCard01
}


def get_model_by_month(month: int):
    model = REPORT_CARD_MODELS.get(month)
    if not model:
        raise HTTPException(status_code=400, detail="Некорректный месяц. Допустимы: 10, 11, 12.")
    return model


@app.get("/api/employee/{month}/{tab}")
async def get_employee_by_tab_and_month(
        month: int = Path(..., ge=10, le=12),
        tab: str = Path(..., min_length=1)
):
    """Получить сотрудника по табельному номеру и месяцу (10=окт, 11=ноя, 12=дек)"""
    model = get_model_by_month(month)
    try:
        emp = model.get(model.tab == tab)
        return {
            "id": emp.id,
            "tab": emp.tab,
            "fio": emp.fio,
            "ksp": emp.ksp,
            "name": emp.name,
            "category": emp.category,
            "profession": emp.profession,
            "status": emp.status,
            "abbreviation": emp.abbreviation,
            "grade": emp.grade,
            "salary": emp.salary,
            "days": json.loads(emp.days)
        }
    except model.DoesNotExist:
        raise HTTPException(status_code=404,
                            detail=f"Сотрудник с табельным номером {tab} не найден в {month} мес. 2025")


@app.put("/api/employee/{month}/{tab}")
async def update_employee_days_by_month(
        month: int = Path(..., ge=10, le=12),
        tab: str = Path(..., min_length=1),
        request: Request = None
):
    """Обновить график сотрудника за указанный месяц"""
    model = get_model_by_month(month)
    try:
        emp = model.get(model.tab == tab)
    except model.DoesNotExist:
        raise HTTPException(status_code=404, detail=f"Сотрудник не найден в {month} мес. 2025")

    try:
        data = await request.json()
        new_days = data.get("days")

        if not isinstance(new_days, list):
            raise HTTPException(status_code=400, detail="Поле 'days' должно быть списком")

        # Проверка количества дней
        days_in_month = {10: 31, 11: 30, 12: 31}
        expected_days = days_in_month[month]
        if len(new_days) != expected_days:
            raise HTTPException(
                status_code=400,
                detail=f"{month} месяц 2025 имеет {expected_days} дней. Передано: {len(new_days)}"
            )

        emp.days = json.dumps(new_days, ensure_ascii=False)
        emp.date_change = datetime.now()
        emp.save()

        return {"status": "ok", "message": f"График за {month} месяц успешно обновлён"}
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Некорректный JSON")
    except Exception as e:
        logger.exception(e)
        raise HTTPException(status_code=500, detail="Ошибка при сохранении")


@app.get("/api/calendar/{year}/{month}")
async def get_calendar_structure(year: int, month: int):
    """Возвращает структуру календаря: смещение + количество дней"""
    if not (1 <= month <= 12):
        raise HTTPException(status_code=400, detail="Некорректный месяц")

    # Количество дней в месяце
    days_in_month = calendar.monthrange(year, month)[1]

    # Первый день месяца: 0=Пн, ..., 6=Вс (как в вашем JS)
    first_weekday = calendar.weekday(year, month, 1)  # Пн=0, Вс=6

    return {
        "year": year,
        "month": month,
        "days_in_month": days_in_month,
        "offset": first_weekday,  # сколько пустых ячеек до 1-го числа
        "weekdays": ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]
    }


@app.post("/add_employee/")
async def add_employee(
        request: Request,
        service_number: Annotated[str, Form()],
        person: Annotated[str, Form()],
        salary: Annotated[str, Form()],
        status: Annotated[str, Form()],
        profession: Annotated[str, Form()],
        admission_date: Annotated[str, Form()],  # или dismissal_date, если нужно
        ksp: Annotated[str, Form()],
):
    """добавление сотрудника в базу данных"""
    try:
        # Преобразуем дату поступления (если используется)
        admission_date = datetime.strptime(admission_date, "%Y-%m-%d").date()
    except ValueError:
        message = "⚠️ Неверный формат даты. Ожидается ГГГГ-ММ-ДД."
        return templates.TemplateResponse("personal_business.html", {"request": request, "message": message})

    try:
        # 1. Добавляем в основную базу
        with db.atomic():
            # Создаём запись в DataStaff
            DataStaff.create(
                service_number=service_number,
                person=person,
                salary=salary,
                status=status,
                profession=profession,
                admission_date=admission_date,
                dismissal_date=None,  # так как это новый сотрудник
                ksp=ksp
            )

        now = datetime.now()

        # 2. Добавляем в ноябрь (ReportCard11), если дата приёма ≤ 30 ноября
        if admission_date <= date(2025, 11, 30):
            days_nov = [""] * 30  # 30 дней в ноябре
            ReportCard11.create(
                ksp=ksp,
                name="",
                category="",
                profession=profession,
                status=status,
                abbreviation="",
                grade="",
                tab=service_number,
                fio=person,
                salary=salary,
                days=json.dumps(days_nov, ensure_ascii=False),
                date_change=now
            )

        # 3. Добавляем в декабрь (ReportCard12), если дата приёма ≤ 31 декабря
        if admission_date <= date(2025, 12, 31):
            days_dec = [""] * 31  # 31 день в декабре
            ReportCard12.create(
                ksp=ksp,
                name="",
                category="",
                profession=profession,
                status=status,
                abbreviation="",
                grade="",
                tab=service_number,
                fio=person,
                salary=salary,
                days=json.dumps(days_dec, ensure_ascii=False),
                date_change=now
            )

        message = f"✅ Сотрудник {person} (таб. №{service_number}) успешно добавлен."
    except IntegrityError:
        message = f"⚠️ Ошибка: сотрудник с табельным номером {service_number} уже существует."
    except Exception as e:
        logger.exception(e)
        message = "⚠️ Произошла ошибка при добавлении сотрудника."

    return templates.TemplateResponse("personal_business.html", {"request": request, "message": message})


@app.get("/")
async def index(request: Request):
    # Передаем контекст в шаблон
    return templates.TemplateResponse("index.html", {"request": request})


@app.get("/admin", response_class=HTMLResponse)
async def admin_panel(request: Request, _username: str = Depends(authenticate_user)):
    """
    Админ-панель (требует авторизации)
    """
    return templates.TemplateResponse("admin/index.html", {"request": request})
