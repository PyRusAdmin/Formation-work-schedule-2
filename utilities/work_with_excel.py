# -*- coding: utf-8 -*-
from openpyxl import load_workbook  # https://openpyxl.readthedocs.io/en/stable/


def get_data_from_excel():
    # Загружаем существующий файл
    wb = load_workbook("data/Списочный_состав.xlsx")
    # Получаем активный лист (тот, который был открыт при сохранении)
    ws = wb.active
    # Читаем данные из ячейки
    print(ws["F6"].value)

    data_list = []

    # Чтение строк с 6 по 189
    for row in ws.iter_rows(min_row=6, max_row=189, values_only=True):
        print(row)
        data_list.append(row)
        print(row[5], row[6], row[3])
    print(data_list)
    return data_list
